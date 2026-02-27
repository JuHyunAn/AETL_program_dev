"""
================================================================================
AETL Template Profile Engine  —  v1.0
================================================================================
사용자 정의 엑셀 양식을 한 번 학습하고, 이후 자동으로 매핑 데이터를 기입합니다.

동작 방식:
  1. [등록] 사용자가 빈 양식 업로드 → AETL이 헤더를 분석 → 사용자가 UI에서 확인/수정
  2. [저장] 확정된 헤더↔필드 매핑을 profile.json + 원본 템플릿 파일로 저장
  3. [적용] 이후 export 시 profile을 로드 → 원본 양식에 데이터를 write-back → 다운로드

핵심 원칙:
  - AI는 초안(헤더 후보 제안)만 담당, 사람이 최종 확정
  - 확정 후에는 100% 결정론적으로 동작 (LLM 없음)
  - 원본 양식 파일을 그대로 사용 (새 파일 생성 X)
================================================================================
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

PROFILE_DIR = Path(__file__).parent / ".template_profiles"

# ─────────────────────────────────────────────────────────────
# AETL 표준 필드 목록  (사용자가 헤더와 연결할 수 있는 필드들)
# ─────────────────────────────────────────────────────────────

AETL_FIELDS: dict[str, str] = {
    # ── 개요 단일값 ──
    "mapping_id":    "매핑 ID",
    "author":        "작성자",
    "created_date":  "작성일",
    "source_table":  "소스 테이블명",
    "target_table":  "타겟 테이블명",
    "load_type":     "적재 유형 (MERGE/INSERT 등)",
    # ── 컬럼 매핑 반복행 ──
    "target_col":    "타겟 컬럼명",
    "source_col":    "소스 컬럼명",
    "transform_rule": "변환 규칙/식",
    "transform_type": "변환 유형 (1:1/집계/파생)",
    "col_description": "컬럼 비고",
    # ── SQL 블록 ──
    "load_sql":      "적재 SQL",
    "validation_sql": "검증 SQL",
    # ── (무시) ──
    "__ignore__":    "이 컬럼은 무시",
}

FIELD_LABELS = {k: v for k, v in AETL_FIELDS.items() if k != "__ignore__"}


# ─────────────────────────────────────────────────────────────
# 1. 템플릿 분석 — 헤더 후보 감지
# ─────────────────────────────────────────────────────────────

def detect_template_structure(file_bytes: bytes) -> dict[str, Any]:
    """
    엑셀 파일을 읽어 시트별 헤더 후보를 반환합니다.
    사용자가 UI에서 헤더↔필드를 확인하는 데 사용합니다.

    Returns:
        {
          sheet_name: {
            "headers": [{"cell_value": str, "row": int, "col": int}],
            "max_row": int,
            "max_col": int,
            "sample_rows": [[...]]   # 2~4행 샘플 데이터
          }
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result: dict[str, Any] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers: list[dict] = []
        header_row = 1

        # 첫 5행에서 가장 많은 값을 가진 행을 헤더 행으로 판단
        best_row, best_count = 1, 0
        for r in range(1, 6):
            count = sum(1 for c in range(1, ws.max_column + 1)
                        if ws.cell(row=r, column=c).value is not None)
            if count > best_count:
                best_count, best_row = count, r
        header_row = best_row

        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is not None and str(val).strip():
                headers.append({
                    "cell_value": str(val).strip(),
                    "row": header_row,
                    "col": col,
                })

        # 샘플 데이터 (헤더 다음 최대 3행)
        sample_rows = []
        for r in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in row_vals):
                sample_rows.append([str(v) if v is not None else "" for v in row_vals])

        result[sheet_name] = {
            "headers":    headers,
            "header_row": header_row,
            "max_row":    ws.max_row,
            "max_col":    ws.max_column,
            "sample_rows": sample_rows,
        }

    return result


def suggest_field_mapping(headers: list[dict]) -> dict[str, str]:
    """
    헤더 텍스트를 기반으로 AETL 필드 매핑을 휴리스틱으로 제안합니다.
    사용자가 수정할 초안을 제공합니다.

    Returns:
        {col_index_str: field_name}  예: {"3": "target_col", "5": "source_col"}
    """
    KEYWORD_MAP: list[tuple[list[str], str]] = [
        (["타겟 컬럼", "target col", "대상 컬럼", "목적 컬럼", "적재 컬럼"], "target_col"),
        (["소스 컬럼", "source col", "원천 컬럼", "원본 컬럼"], "source_col"),
        (["변환", "transform", "변환식", "변환규칙", "매핑규칙", "규칙"], "transform_rule"),
        (["변환유형", "transform type", "매핑유형", "유형"], "transform_type"),
        (["비고", "remark", "설명", "description", "comment"], "col_description"),
        (["매핑id", "mapping id", "매핑번호"], "mapping_id"),
        (["작성자", "author", "담당자"], "author"),
        (["작성일", "created", "작성일자", "일자"], "created_date"),
        (["소스테이블", "source table", "원천테이블"], "source_table"),
        (["타겟테이블", "target table", "대상테이블", "적재테이블"], "target_table"),
        (["적재sql", "load sql", "적재쿼리"], "load_sql"),
        (["검증sql", "validation sql", "검증쿼리"], "validation_sql"),
        (["적재유형", "load type", "적재방식"], "load_type"),
    ]

    result: dict[str, str] = {}
    for h in headers:
        val_lower = h["cell_value"].lower().replace(" ", "").replace("_", "")
        matched = "__ignore__"
        for keywords, field in KEYWORD_MAP:
            if any(kw.replace(" ", "").replace("_", "") in val_lower for kw in keywords):
                matched = field
                break
        result[str(h["col"])] = matched

    return result


# ─────────────────────────────────────────────────────────────
# 2. 프로파일 저장 / 로드
# ─────────────────────────────────────────────────────────────

def save_profile(
    profile_name: str,
    sheet_configs: list[dict],
    template_bytes: bytes,
) -> None:
    """
    Args:
        profile_name:   저장할 프로파일 이름
        sheet_configs:  [
          {
            "sheet_name": str,
            "sheet_type": "overview" | "column_mapping" | "sql_load" | "sql_validation",
            "header_row": int,
            "data_start_row": int,     # column_mapping 타입에서 데이터 시작 행
            "col_field_map": {col_idx(str): field_name},  # 헤더 컬럼 → 필드
            "sql_cell": {"row": int, "col": int},         # sql 타입에서 SQL 기입 위치
          }
        ]
        template_bytes: 원본 엑셀 파일 bytes
    """
    PROFILE_DIR.mkdir(exist_ok=True)

    profile_data = {
        "name":           profile_name,
        "created_at":     datetime.now().isoformat(),
        "sheet_configs":  sheet_configs,
    }

    profile_path  = PROFILE_DIR / f"{profile_name}.json"
    template_path = PROFILE_DIR / f"{profile_name}_template.xlsx"

    with open(profile_path, "w", encoding="utf-8") as f:
        json.dump(profile_data, f, ensure_ascii=False, indent=2)
    with open(template_path, "wb") as f:
        f.write(template_bytes)


def load_profile(profile_name: str) -> tuple[dict | None, bytes]:
    """
    Returns:
        (profile_data, template_bytes)
        profile_data가 None이면 프로파일이 없는 것
    """
    profile_path  = PROFILE_DIR / f"{profile_name}.json"
    template_path = PROFILE_DIR / f"{profile_name}_template.xlsx"

    if not profile_path.exists():
        return None, b""

    with open(profile_path, encoding="utf-8") as f:
        profile_data = json.load(f)

    template_bytes = b""
    if template_path.exists():
        with open(template_path, "rb") as f:
            template_bytes = f.read()

    return profile_data, template_bytes


def list_profiles() -> list[str]:
    """저장된 프로파일 이름 목록을 반환합니다."""
    if not PROFILE_DIR.exists():
        return []
    return sorted(p.stem for p in PROFILE_DIR.glob("*.json"))


def delete_profile(profile_name: str) -> bool:
    """프로파일과 원본 템플릿 파일을 삭제합니다."""
    deleted = False
    for ext in (".json", "_template.xlsx"):
        p = PROFILE_DIR / f"{profile_name}{ext}"
        if p.exists():
            p.unlink()
            deleted = True
    return deleted


# ─────────────────────────────────────────────────────────────
# 3. Write-back — 원본 양식에 데이터 기입
# ─────────────────────────────────────────────────────────────

def apply_profile(
    profile_data: dict,
    template_bytes: bytes,
    mapping_result: dict,
) -> bytes:
    """
    저장된 프로파일에 따라 원본 엑셀 양식에 매핑 데이터를 기입합니다.

    Args:
        profile_data:   load_profile()로 가져온 프로파일 JSON
        template_bytes: 원본 빈 양식 bytes
        mapping_result: {
            "mapping_id":      str,
            "author":          str,
            "load_type":       str,
            "source_meta":     dict,
            "target_meta":     dict,
            "column_mappings": list[dict],
            "load_sql":        str,
            "validation_sqls": list[dict],
        }

    Returns:
        기입 완료된 xlsx bytes
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

    # 단일값 조회 테이블
    single_values: dict[str, str] = {
        "mapping_id":   mapping_result.get("mapping_id", ""),
        "author":       mapping_result.get("author", "AETL"),
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "source_table": mapping_result.get("source_meta", {}).get("table_name", ""),
        "target_table": mapping_result.get("target_meta", {}).get("table_name", ""),
        "load_type":    mapping_result.get("load_type", "MERGE"),
    }

    column_mappings  = mapping_result.get("column_mappings", [])
    load_sql         = mapping_result.get("load_sql", "")
    validation_sqls  = mapping_result.get("validation_sqls", [])

    for cfg in profile_data.get("sheet_configs", []):
        sheet_name = cfg.get("sheet_name", "")
        if sheet_name not in wb.sheetnames:
            continue

        ws         = wb[sheet_name]
        sheet_type = cfg.get("sheet_type", "overview")
        col_field  = {int(k): v for k, v in cfg.get("col_field_map", {}).items()}

        if sheet_type == "overview":
            _writeback_overview(ws, col_field, single_values, cfg)

        elif sheet_type == "column_mapping":
            _writeback_column_mapping(ws, col_field, column_mappings, cfg)

        elif sheet_type == "sql_load":
            pos = cfg.get("sql_cell", {"row": 2, "col": 1})
            ws.cell(row=pos["row"], column=pos["col"], value=load_sql)

        elif sheet_type == "sql_validation":
            pos = cfg.get("sql_cell", {"row": 2, "col": 1})
            combined = "\n\n".join(
                f"-- {s.get('name','')}\n{s.get('sql','')}"
                for s in validation_sqls
            )
            ws.cell(row=pos["row"], column=pos["col"], value=combined)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _writeback_overview(ws, col_field: dict, single_values: dict, cfg: dict):
    """개요형 시트: 헤더 옆(오른쪽) 또는 아래 셀에 단일값 기입"""
    header_row = cfg.get("header_row", 1)

    for col_idx, field_name in col_field.items():
        if field_name in ("__ignore__", "") or field_name not in single_values:
            continue
        value = single_values[field_name]
        # 헤더가 세로 배치(A열 레이블, B열 값)인지 가로 배치인지 판단
        # 헤더 행의 컬럼이 2개 이하면 세로(레이블-값 형) → 같은 행 오른쪽 열에 기입
        # 그 외(가로 테이블 헤더)면 다음 행에 기입
        headers_in_row = sum(1 for c, f in col_field.items() if f and f != "__ignore__")
        if headers_in_row <= 3:
            # 세로형: 레이블 오른쪽에 값 기입
            ws.cell(row=header_row + (col_idx - 1), column=col_idx + 1, value=value)
        else:
            # 가로형: 헤더 바로 아래 행에 기입
            ws.cell(row=header_row + 1, column=col_idx, value=value)


def _writeback_column_mapping(ws, col_field: dict, column_mappings: list[dict], cfg: dict):
    """컬럼 매핑형 시트: data_start_row부터 한 행씩 반복 기입"""
    data_start = cfg.get("data_start_row", cfg.get("header_row", 1) + 1)

    row_field_value: dict[str, str]
    for i, mapping in enumerate(column_mappings):
        row = data_start + i
        row_field_value = {
            "target_col":      mapping.get("target_col", ""),
            "source_col":      mapping.get("source_col", ""),
            "transform_rule":  mapping.get("transform", ""),
            "transform_type":  mapping.get("transform_type", "1:1"),
            "col_description": mapping.get("description", ""),
        }
        for col_idx, field_name in col_field.items():
            if field_name in row_field_value:
                ws.cell(row=row, column=col_idx, value=row_field_value[field_name])
