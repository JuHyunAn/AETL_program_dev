"""
================================================================================
AETL Export Engine  —  v2.1
================================================================================
산출물 자동 생성:
  1. 매핑정의서 Excel (openpyxl)
  2. DDL Script (db_type별 방언)
  3. 검증 리포트 Excel
================================================================================
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# 공통 스타일
# ─────────────────────────────────────────────────────────────

_BLUE   = "0070C0"
_DARK   = "1F2D3D"
_LIGHT  = "EEF1F6"
_GREEN  = "16A34A"
_RED    = "DC2626"
_WHITE  = "FFFFFF"

def _header_style(bold=True, bg=_DARK, fg=_WHITE, size=10):
    return {
        "font":      Font(bold=bold, color=fg, size=size),
        "fill":      PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }

def _cell_style(bg=_LIGHT, fg="1A202C", size=10, bold=False, align="left"):
    return {
        "font":      Font(bold=bold, color=fg, size=size),
        "fill":      PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }

def _thin_border():
    s = Side(border_style="thin", color="D0D9E4")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply(cell, style: dict):
    for k, v in style.items():
        setattr(cell, k, v)

def _write_header_row(ws, headers: list[str], row: int = 1):
    st = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        _apply(cell, st)

def _write_data_row(ws, values: list, row: int, alt: bool = False):
    bg = "F5F7FA" if alt else _WHITE
    st = _cell_style(bg=bg)
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        _apply(cell, st)


# ─────────────────────────────────────────────────────────────
# 1. 매핑정의서 Excel 생성
# ─────────────────────────────────────────────────────────────

def generate_mapping_excel(
    source_meta: dict,
    target_meta: dict,
    column_mappings: list[dict],
    load_sql: str = "",
    validation_sqls: list[dict] | None = None,
    mapping_id: str = "",
    author: str = "AETL",
) -> bytes:
    """
    매핑정의서 Excel을 생성하고 bytes로 반환합니다.

    Args:
        source_meta:      {"table_name": str, "columns": [...], "pk_columns": [...]}
        target_meta:      동일 구조
        column_mappings:  [{"source_col": str, "target_col": str, "transform_rule": str,
                            "transform_type": str, "description": str}]
        load_sql:         적재 SQL 문자열
        validation_sqls:  [{"name": str, "sql": str, "expected": str}]
        mapping_id:       매핑 ID
        author:           작성자

    Returns:
        xlsx 파일 bytes
    """
    wb = openpyxl.Workbook()

    # ── validation_sqls 정규화: dict[str, dict] → list[dict] 자동 변환 ──
    if validation_sqls is None:
        validation_sqls = []
    elif isinstance(validation_sqls, dict):
        normalized: list[dict] = []
        for key, val in validation_sqls.items():
            if isinstance(val, dict):
                normalized.append({
                    "name": val.get("name", val.get("rule_name", key)),
                    "sql":  val.get("sql", ""),
                    "expected": val.get("expected", ""),
                })
            else:
                normalized.append({"name": key, "sql": str(val), "expected": ""})
        validation_sqls = normalized

    # ── Sheet 1: 개요 ──────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "개요"
    ws_ov.column_dimensions["A"].width = 20
    ws_ov.column_dimensions["B"].width = 35

    overview_rows = [
        ("매핑 ID",       mapping_id or f"MAP_{datetime.now():%Y%m%d%H%M%S}"),
        ("작성일",        datetime.now().strftime("%Y-%m-%d")),
        ("작성자",        author),
        ("소스 테이블",   source_meta.get("table_name", "")),
        ("타겟 테이블",   target_meta.get("table_name", "")),
        ("소스 컬럼 수",  len(source_meta.get("columns", []))),
        ("타겟 컬럼 수",  len(target_meta.get("columns", []))),
        ("매핑 컬럼 수",  len(column_mappings)),
    ]
    h_st = _header_style()
    d_st = _cell_style()
    for r, (label, val) in enumerate(overview_rows, 2):
        c1 = ws_ov.cell(row=r, column=1, value=label)
        c2 = ws_ov.cell(row=r, column=2, value=str(val))
        _apply(c1, h_st); _apply(c2, d_st)

    # 타이틀
    title_cell = ws_ov.cell(row=1, column=1, value="ETL 매핑정의서")
    title_cell.font = Font(bold=True, size=14, color=_DARK)
    ws_ov.merge_cells("A1:B1")

    # ── Sheet 2: 소스 테이블 정보 ─────────────────────────────
    ws_src = wb.create_sheet("소스 테이블")
    _fill_table_info_sheet(ws_src, source_meta)

    # ── Sheet 3: 타겟 테이블 정보 ─────────────────────────────
    ws_tgt = wb.create_sheet("타겟 테이블")
    _fill_table_info_sheet(ws_tgt, target_meta)

    # ── Sheet 4: 컬럼 매핑 ────────────────────────────────────
    ws_map = wb.create_sheet("컬럼 매핑")
    headers = ["No", "타겟 컬럼", "타겟 타입", "소스 컬럼", "소스 타입",
               "변환 규칙", "변환 유형", "비고"]
    _write_header_row(ws_map, headers)
    for col_w, w in zip("ABCDEFGH", [5, 20, 12, 20, 12, 35, 12, 20]):
        ws_map.column_dimensions[col_w].width = w

    # 소스/타겟 컬럼 타입 맵 구성
    src_type_map = {c["name"]: c.get("type", "") for c in source_meta.get("columns", [])}
    tgt_type_map = {c["name"]: c.get("type", "") for c in target_meta.get("columns", [])}

    for i, m in enumerate(column_mappings, 1):
        vals = [
            i,
            m.get("target_col", ""),
            tgt_type_map.get(m.get("target_col", ""), ""),
            m.get("source_col", ""),
            src_type_map.get(m.get("source_col", ""), ""),
            m.get("transform_rule", ""),
            m.get("transform_type", "1:1"),
            m.get("description", ""),
        ]
        _write_data_row(ws_map, vals, i + 1, alt=i % 2 == 0)

    # ── Sheet 5: 적재 SQL ─────────────────────────────────────
    ws_sql = wb.create_sheet("적재 SQL")
    ws_sql.column_dimensions["A"].width = 120
    ws_sql.cell(row=1, column=1, value="-- 적재 SQL (AETL 자동 생성)").font = Font(bold=True, color=_BLUE)
    ws_sql.cell(row=2, column=1, value=load_sql or "-- (적재 SQL 없음)")
    ws_sql.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_sql.row_dimensions[2].height = max(60, load_sql.count("\n") * 14)

    # ── Sheet 6: 검증 SQL ─────────────────────────────────────
    ws_val = wb.create_sheet("검증 SQL")
    _write_header_row(ws_val, ["No", "검증명", "검증 SQL", "기대 결과"])
    for col_w, w in zip("ABCD", [5, 20, 80, 20]):
        ws_val.column_dimensions[col_w].width = w
    for i, v in enumerate(validation_sqls, 1):
        _write_data_row(ws_val, [
            i,
            v.get("name", ""),
            v.get("sql", ""),
            v.get("expected", ""),
        ], i + 1, alt=i % 2 == 0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_table_info_sheet(ws, meta: dict):
    headers = ["No", "스키마", "테이블명", "컬럼명", "데이터타입", "길이", "NULL여부", "PK", "설명"]
    _write_header_row(ws, headers)
    for col_w, w in zip("ABCDEFGHI", [5, 15, 20, 25, 15, 8, 10, 5, 25]):
        ws.column_dimensions[col_w].width = w

    table_name = meta.get("table_name", "")
    pk_cols = set(meta.get("pk_columns", []))
    for i, col in enumerate(meta.get("columns", []), 1):
        cname = col.get("name", col.get("column_name", ""))
        _write_data_row(ws, [
            i, "", table_name, cname,
            col.get("type", col.get("data_type", "")),
            col.get("length", ""),
            "N" if col.get("nullable", True) else "Y",
            "PK" if cname in pk_cols else "",
            col.get("description", ""),
        ], i + 1, alt=i % 2 == 0)


# ─────────────────────────────────────────────────────────────
# 2. DDL Script 생성
# ─────────────────────────────────────────────────────────────

_TYPE_MAP = {
    # Oracle → target
    "oracle": {
        "oracle": {},  # 변환 없음
        "mariadb": {
            "VARCHAR2": "VARCHAR", "NUMBER": "DECIMAL", "DATE": "DATETIME",
            "CLOB": "LONGTEXT", "BLOB": "LONGBLOB", "CHAR": "CHAR",
        },
        "postgresql": {
            "VARCHAR2": "VARCHAR", "NUMBER": "NUMERIC", "DATE": "TIMESTAMP",
            "CLOB": "TEXT", "BLOB": "BYTEA", "CHAR": "CHAR",
        },
    }
}


def generate_ddl(table_meta: dict, db_type: str = "oracle") -> str:
    """
    테이블 메타데이터로부터 CREATE TABLE DDL을 생성합니다.

    Args:
        table_meta: {"table_name": str, "columns": [...], "pk_columns": [...]}
        db_type:    oracle | mariadb | postgresql

    Returns:
        DDL 문자열
    """
    table = table_meta.get("table_name", "TABLE_NAME")
    columns = table_meta.get("columns", [])
    pk_cols = table_meta.get("pk_columns", [])

    lines = []
    if db_type == "oracle":
        lines.append(f'CREATE TABLE "{table}" (')
    elif db_type == "mariadb":
        lines.append(f"CREATE TABLE `{table}` (")
    else:
        lines.append(f'CREATE TABLE "{table}" (')

    col_defs = []
    for col in columns:
        cname = col.get("name", col.get("column_name", "COL"))
        ctype = col.get("type", col.get("data_type", "VARCHAR2(200)"))
        nullable = col.get("nullable", True)
        default = col.get("default", None)

        if db_type == "oracle":
            null_str = "" if nullable else " NOT NULL"
            def_str  = f" DEFAULT {default}" if default else ""
            col_defs.append(f'    "{cname}" {ctype}{def_str}{null_str}')
        elif db_type == "mariadb":
            mapped = _TYPE_MAP.get("oracle", {}).get("mariadb", {}).get(
                ctype.upper().split("(")[0], ctype)
            null_str = "" if nullable else " NOT NULL"
            def_str  = f" DEFAULT {default}" if default else ""
            col_defs.append(f"    `{cname}` {mapped}{def_str}{null_str}")
        else:
            mapped = _TYPE_MAP.get("oracle", {}).get("postgresql", {}).get(
                ctype.upper().split("(")[0], ctype)
            null_str = "" if nullable else " NOT NULL"
            def_str  = f" DEFAULT {default}" if default else ""
            col_defs.append(f'    "{cname}" {mapped}{def_str}{null_str}')

    if pk_cols:
        if db_type == "oracle":
            pk_str = ", ".join(f'"{c}"' for c in pk_cols)
            col_defs.append(f"    CONSTRAINT PK_{table} PRIMARY KEY ({pk_str})")
        elif db_type == "mariadb":
            pk_str = ", ".join(f"`{c}`" for c in pk_cols)
            col_defs.append(f"    PRIMARY KEY ({pk_str})")
        else:
            pk_str = ", ".join(f'"{c}"' for c in pk_cols)
            col_defs.append(f"    PRIMARY KEY ({pk_str})")

    lines.append(",\n".join(col_defs))
    lines.append(");")
    return "\n".join(lines)


def generate_merge_sql(
    source_meta: dict, target_meta: dict,
    column_mappings: list[dict], db_type: str = "oracle"
) -> str:
    """소스 → 타겟 MERGE/UPSERT 적재 SQL 생성"""
    src = source_meta.get("table_name", "SOURCE")
    tgt = target_meta.get("table_name", "TARGET")
    pk_cols = target_meta.get("pk_columns", [])

    if not pk_cols:
        return "-- ⚠ PK 컬럼이 정의되지 않아 MERGE SQL을 생성할 수 없습니다."

    if db_type == "oracle":
        on_cond = " AND ".join(f'tgt."{p}" = src."{p}"' for p in pk_cols)
        set_cols = [m for m in column_mappings if m.get("target_col") not in pk_cols]
        update_str = ",\n        ".join(
            f'tgt."{m["target_col"]}" = src."{m.get("source_col", m["target_col"])}"'
            for m in set_cols
        )
        insert_tgt = ", ".join(f'"{m["target_col"]}"' for m in column_mappings)
        insert_src = ", ".join(f'src."{m.get("source_col", m["target_col"])}"' for m in column_mappings)
        return f"""MERGE INTO "{tgt}" tgt
USING (SELECT * FROM "{src}") src
ON ({on_cond})
WHEN MATCHED THEN
    UPDATE SET
        {update_str}
WHEN NOT MATCHED THEN
    INSERT ({insert_tgt})
    VALUES ({insert_src});"""

    elif db_type == "mariadb":
        cols_tgt = ", ".join(f'`{m["target_col"]}`' for m in column_mappings)
        cols_src = ", ".join(f'`{m.get("source_col", m["target_col"])}`' for m in column_mappings)
        update_str = ", ".join(
            f'`{m["target_col"]}` = VALUES(`{m["target_col"]}`)'
            for m in column_mappings
        )
        return f"""INSERT INTO `{tgt}` ({cols_tgt})
SELECT {cols_src}
FROM `{src}`
ON DUPLICATE KEY UPDATE
    {update_str};"""

    else:  # postgresql
        cols_tgt = ", ".join(f'"{m["target_col"]}"' for m in column_mappings)
        cols_src = ", ".join(f'"{m.get("source_col", m["target_col"])}"' for m in column_mappings)
        pk_str = ", ".join(f'"{p}"' for p in pk_cols)
        update_str = ", ".join(
            f'"{m["target_col"]}" = EXCLUDED."{m["target_col"]}"'
            for m in column_mappings
        )
        return f"""INSERT INTO "{tgt}" ({cols_tgt})
SELECT {cols_src}
FROM "{src}"
ON CONFLICT ({pk_str})
DO UPDATE SET
    {update_str};"""


# ─────────────────────────────────────────────────────────────
# 3. 검증 리포트 Excel 생성
# ─────────────────────────────────────────────────────────────

def generate_validation_report(
    run_results: list[dict],
    mapping_id: str = "",
    source_table: str = "",
    target_table: str = "",
) -> bytes:
    """
    검증 실행 결과를 Excel 리포트로 생성합니다.

    Args:
        run_results: [
          {
            "name": str,           # 검증명
            "sql": str,            # 실행 SQL
            "ok": bool,            # PASS/FAIL
            "result": Any,         # 실행 결과 요약
            "elapsed_sec": float,
            "error": str | None,
          }
        ]

    Returns:
        xlsx bytes
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: 요약 ─────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "검증 요약"
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 40

    # 헤더
    title = ws_sum.cell(row=1, column=1, value="ETL 검증 리포트")
    title.font = Font(bold=True, size=14, color=_DARK)
    ws_sum.merge_cells("A1:B1")

    # ── run_results 키 정규화: rule_name/status 형식도 허용 ──
    def _normalize_result(r: dict) -> dict:
        ok_val = r.get("ok")
        if ok_val is None:
            status_str = str(r.get("status", "FAIL")).upper()
            ok_val = status_str == "PASS"
        return {
            "name":        r.get("name") or r.get("rule_name", ""),
            "sql":         r.get("sql", ""),
            "ok":          bool(ok_val),
            "result":      r.get("result", r.get("actual_value", "")),
            "elapsed_sec": r.get("elapsed_sec", ""),
            "error":       r.get("error", ""),
        }
    run_results = [_normalize_result(r) for r in run_results]

    total   = len(run_results)
    passed  = sum(1 for r in run_results if r.get("ok"))
    failed  = total - passed
    pass_rate = f"{passed/total*100:.1f}%" if total else "N/A"

    summary_rows = [
        ("매핑 ID",       mapping_id),
        ("소스 테이블",   source_table),
        ("타겟 테이블",   target_table),
        ("실행 일시",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("전체 검증 수",  total),
        ("PASS",          passed),
        ("FAIL",          failed),
        ("통과율",        pass_rate),
    ]

    h_st = _header_style()
    for r, (label, val) in enumerate(summary_rows, 3):
        c1 = ws_sum.cell(row=r, column=1, value=label)
        c2 = ws_sum.cell(row=r, column=2, value=str(val))
        _apply(c1, h_st)
        _apply(c2, _cell_style(bg="F0F6FF" if r % 2 == 0 else _WHITE))

    # ── Sheet 2: 상세 결과 ────────────────────────────────────
    ws_det = wb.create_sheet("상세 결과")
    headers = ["No", "검증명", "PASS/FAIL", "실행 결과", "소요 시간(초)", "오류 메시지"]
    _write_header_row(ws_det, headers)
    for col_w, w in zip("ABCDEF", [5, 25, 12, 50, 12, 40]):
        ws_det.column_dimensions[col_w].width = w

    for i, res in enumerate(run_results, 1):
        ok     = res.get("ok", False)
        status = "PASS" if ok else "FAIL"
        vals   = [
            i,
            res.get("name", ""),
            status,
            str(res.get("result", ""))[:200],
            res.get("elapsed_sec", ""),
            res.get("error", "") or "",
        ]
        _write_data_row(ws_det, vals, i + 1, alt=i % 2 == 0)

        # FAIL 행 빨간 강조
        if not ok:
            for col in range(1, 7):
                ws_det.cell(row=i + 1, column=col).fill = PatternFill("solid", fgColor="FEE2E2")

    # ── Sheet 3: SQL 목록 ─────────────────────────────────────
    ws_sql = wb.create_sheet("실행 SQL")
    headers_sql = ["No", "검증명", "실행 SQL"]
    _write_header_row(ws_sql, headers_sql)
    ws_sql.column_dimensions["A"].width = 5
    ws_sql.column_dimensions["B"].width = 25
    ws_sql.column_dimensions["C"].width = 100

    for i, res in enumerate(run_results, 1):
        _write_data_row(ws_sql, [i, res.get("name", ""), res.get("sql", "")], i + 1)
        ws_sql.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_sql.row_dimensions[i + 1].height = max(30, res.get("sql", "").count("\n") * 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
